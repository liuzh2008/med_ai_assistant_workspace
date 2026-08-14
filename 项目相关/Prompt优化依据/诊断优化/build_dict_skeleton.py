# -*- coding: utf-8 -*-
"""
诊断编码词典构建器（第1步：名称层骨架）

输入：
  - diagnosis_analysis_records.txt   (AI诊断分析记录)
  - current_diagnoses.txt            (医生目前诊断)
输出：
  - 诊断编码词典_骨架.json             (2577 AI名 → 标准名桥接 + ICD编码待回填位)
  - 诊断编码词典_人工审核清单.xlsx      (686 个无法自动桥接的 AI 名，供人工编码/复核)
  - 医生标准名清单.json               (671 个医生标准名，icd10_code 待生产回填)

后续（第2步）：生产库重新导出含 ICD10Code 的 current_diagnoses.txt 后，
运行 fill_icd_codes.py 回填 icd10_code，得到最终词典。

用法：python build_dict_skeleton.py
"""
import importlib.util
import json
import os
from collections import Counter, defaultdict
from datetime import datetime

BASE = os.path.dirname(os.path.abspath(__file__))
spec = importlib.util.spec_from_file_location('pag', os.path.join(BASE, 'parse_and_generate.py'))
pag = importlib.util.module_from_spec(spec)
spec.loader.exec_module(pag)

SIM_MODIFIED = pag.SIM_MODIFIED  # 0.45

# ---------------------------------------------------------------- 加载

def load_records():
    return pag.load_records(os.path.join(BASE, 'diagnosis_analysis_records.txt'))

def load_diagnoses():
    by_patient = defaultdict(list)
    with open(os.path.join(BASE, 'current_diagnoses.txt'), encoding='utf-8') as f:
        for line in f:
            line = line.rstrip('\r\n')
            if not line:
                continue
            parts = line.split(pag.SEP, maxsplit=6)
            if len(parts) < 6:
                continue
            by_patient[parts[0].strip()].append({'text': pag.unescape(parts[1].strip()),
                                                 'is_primary': parts[2].strip() == '1'})
    return by_patient

# ---------------------------------------------------------------- 统计

def collect_ai_names(records):
    """AI 诊断名 -> 频次"""
    cnt = Counter()
    for rec in records:
        blocks, _ = pag.parse_ai_diagnoses(rec['content'])
        for b in blocks:
            if b['name']:
                cnt[b['name'].strip()] += 1
    return cnt

def collect_doc_names(by_patient):
    cnt = Counter()
    for ds in by_patient.values():
        for d in ds:
            cnt[d['text'].strip()] += 1
    return cnt

# ---------------------------------------------------------------- 桥接

SIM_HIGH = 0.80  # 高置信度：核心疾病一致、仅修饰语差异（可自动采信映射）

def top_candidates(ai_name, doc_names, topn=3):
    """返回 (best, candidates)；candidates 为 (医生名, sim, 频次) 降序"""
    scored = []
    for dn, dcnt in doc_names.items():
        s = pag.levenshtein_sim(ai_name, dn)
        if s >= SIM_MODIFIED - 0.05:
            scored.append((dn, round(s, 3), dcnt))
    scored.sort(key=lambda x: -x[1])
    if not scored:
        return None, []
    best = scored[0]
    return best, scored[:topn]

# ---------------------------------------------------------------- 输出

def gen_json(ai_names, doc_names):
    entries = []
    n_auto_high, n_auto_low, n_manual = 0, 0, 0
    for name, freq in ai_names.most_common():
        best, cands = top_candidates(name, doc_names)
        if best and best[1] >= 0.999:
            status = 'adopted'
            confidence = 'high'
            n_auto_high += 1
        elif best and best[1] >= SIM_HIGH:
            status = 'modified-high'
            confidence = 'high'
            n_auto_high += 1
        elif best and best[1] >= SIM_MODIFIED:
            status = 'modified-low'
            confidence = 'low'   # 0.45~0.8：存在"张冠李戴"风险（如二尖瓣↔三尖瓣），须人工确认
            n_auto_low += 1
        else:
            status = 'manual'
            confidence = 'low'
            n_manual += 1
        entries.append({
            'ai_name': name,
            'freq': freq,
            'standard_name': best[0] if best else '',
            'bridge_sim': best[1] if best else 0.0,
            'bridge_status': status,
            'confidence': confidence,
            'candidates': [{'name': c[0], 'sim': c[1], 'freq': c[2]} for c in cands],
            'icd10_code': '',
            'need_review': status in ('modified-low', 'manual'),
        })
    payload = {
        'schema_version': '1.0',
        'purpose': 'AI诊断名 → ICD-10 标准诊断名/编码 对照词典（Prompt模板优化 + 分析工具判定用）',
        'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M'),
        'stats': {'ai_names': len(entries), 'auto_bridged_high': n_auto_high,
                  'auto_bridged_low': n_auto_low, 'manual_pending': n_manual,
                  'need_review': n_auto_low + n_manual,
                  'ai_items_total': sum(ai_names.values())},
        'entries': entries,
    }
    path = os.path.join(BASE, '诊断编码词典_骨架.json')
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(payload, f, ensure_ascii=False, indent=1)
    print('[OK] %s  (high=%d low=%d manual=%d)' % (path, n_auto_high, n_auto_low, n_manual))
    return payload

def gen_review_xlsx(entries):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = '人工审核清单'
    headers = ['AI诊断名', '频次', '状态', '最接近医生名', 'sim', '候选2', 'sim2',
               '候选3', 'sim3', '建议标准名(人工填)', 'ICD-10编码(人工填)', '备注']
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill('solid', fgColor='DDEBF7')

    manual = [e for e in entries if e['bridge_status'] == 'manual']
    low = [e for e in entries if e['bridge_status'] == 'modified-low']
    # 审核对象：无法桥接 + 低置信度桥接（0.45~0.8），先高频
    review = sorted(manual + low, key=lambda e: -e['freq'])
    for e in review:
        cds = e['candidates'] + [{'name': '', 'sim': '', 'freq': ''}] * 3
        ws.append([e['ai_name'], e['freq'], e['bridge_status'], cds[0]['name'], cds[0]['sim'],
                   cds[1]['name'], cds[1]['sim'], cds[2]['name'], cds[2]['sim'], '', '', ''])
    ws.freeze_panes = 'A2'
    widths = [38, 8, 10, 38, 7, 38, 7, 38, 7, 38, 16, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    ws2 = wb.create_sheet('医生标准名待回填')
    ws2.append(['医生标准名', '频次', 'ICD-10编码(生产导出后回填)'])
    for c in ws2[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill('solid', fgColor='E2EFDA')
    doc_names = _last_doc_names
    for name, cnt in doc_names.most_common():
        ws2.append([name, cnt, ''])
    for i, w in enumerate([50, 8, 30], 1):
        ws2.column_dimensions[chr(64 + i)].width = w

    path = os.path.join(BASE, '诊断编码词典_人工审核清单.xlsx')
    wb.save(path)
    print('[OK] %s  (待审核 %d 条)' % (path, len(review)))

def gen_doc_list_json(doc_names):
    payload = {
        'schema_version': '1.0',
        'purpose': '医生标准诊断名清单（生产库 DIAGNOSIS.ICD10Code 回填后即为标准名→编码权威映射）',
        'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M'),
        'entries': [{'standard_name': n, 'freq': c, 'icd10_code': ''}
                    for n, c in doc_names.most_common()],
    }
    path = os.path.join(BASE, '医生标准名清单.json')
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(payload, f, ensure_ascii=False, indent=1)
    print('[OK] %s  (%d 个医生标准名)' % (path, len(doc_names)))

def main():
    print('读取诊断分析记录...')
    records = load_records()
    print('读取目前诊断...')
    by_patient = load_diagnoses()
    ai_names = collect_ai_names(records)
    doc_names = collect_doc_names(by_patient)
    global _last_doc_names
    _last_doc_names = doc_names
    print('AI 诊断名 %d 个 / 医生标准名 %d 个' % (len(ai_names), len(doc_names)))
    payload = gen_json(ai_names, doc_names)
    gen_review_xlsx(payload['entries'])
    gen_doc_list_json(doc_names)
    print('完成。下一步：生产库重新导出含 ICD10Code 的 current_diagnoses.txt，运行 fill_icd_codes.py 回填编码。')

if __name__ == '__main__':
    main()
