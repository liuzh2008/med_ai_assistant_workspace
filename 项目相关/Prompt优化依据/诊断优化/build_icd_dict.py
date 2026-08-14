# -*- coding: utf-8 -*-
"""
诊断编码词典构建器（第3版：全量 ICD-10 国临版2.0 + 同义词归一化 + 算法筛查 + 人工校正表）

数据源：
  - 国家临床版2.0疾病诊断编码（ICD-10）.xlsx   （35863 条：编码 + 名称，含扩展码）
  - diagnosis_analysis_records.txt             （AI 诊断名全集 + 频次）
  - current_diagnoses.txt / 医生标准名清单.json  （医生标准名锚点）

产出（全部 UTF-8）：
  - ICD库_国临版2.0.json         全量 ICD 库（code -> name 主/别名索引），可复用
  - 诊断编码词典.json            最终词典：AI 名 → 标准名 + ICD 编码 + 置信度/来源/备注
  - 诊断编码词典_统计报告.md     匹配率统计 + Top 未映射清单 + Prompt 注入清单
  - 诊断编码词典_未映射清单.xlsx  算法无法映射的条目（人工补码用）

用法：
  python build_icd_dict.py                 # 复用现有 ICD库_国临版2.0.json
  python build_icd_dict.py <国临版2.0.xlsx路径>   # 首次：从 xlsx 导出库
"""
import json
import os
import re
import sys
from collections import Counter
from datetime import datetime

BASE = os.path.dirname(os.path.abspath(__file__))

# ================================================================ 归一化

CN_NUM = {'一': '1', '二': '2', '三': '3', '四': '4', '五': '5',
          '六': '6', '七': '7', '八': '8', '九': '9'}
ROMAN = {'Ⅰ': '1', 'Ⅱ': '2', 'Ⅲ': '3', 'Ⅳ': '4', 'Ⅴ': '5'}

# 同义词/术语归一化（先于 Levenshtein，双向生效：查询名与 ICD 名都过此函数）
SYNONYMS = [
    ('主动脉瓣反流', '主动脉瓣关闭不全'),
    ('三尖瓣反流', '三尖瓣关闭不全'),
    ('肺动脉瓣反流', '肺动脉瓣关闭不全'),
    ('左心房增大', '左房增大'), ('右心房增大', '右房增大'),
    ('心房增大', '房增大'),
    ('左房增大', '左房扩大'), ('右房增大', '右房扩大'), ('双房增大', '双房扩大'),
    ('主动脉增宽', '主动脉扩张'),
    ('血脂异常', '高脂血症'),
    ('混合型', '混合性'),
    ('室性早搏', '室性期前收缩'),
    ('房性早搏', '房性期前收缩'),
    ('室上性早搏', '室上性期前收缩'),
    ('肾功能异常', '肾功能不全'),
    ('肝功能异常', '肝功能不全'),
    ('糖耐量异常', '糖耐量受损'),
]
# 修饰词（用于命中优选/包含匹配惩罚；条目名含而查询名不含时扣分）
MODIFIERS = ['风湿性', '非风湿性', '先天性', '后天性', '特发性', '原发性', '继发性',
             '药物性', '恶性', '良性', '家族性', '老年性', '妊娠', '产后', '新生儿',
             '急性', '慢性', '轻度', '中度', '重度', '极重度', '难治性', '免疫性',
             '获得性', '并发', '合并', '伴', '其他', '未特指', '特指', '巨大',
             '感染性', '缺血性', '出血性', '梗阻性', '扩张性', '肥厚性', '限制性']


def norm(s):
    """统一口径归一化：同义词替换 + 去标点/空白 + 小写 + 数字统一"""
    s = s or ''
    for k, v in SYNONYMS:
        s = s.replace(k, v)
    # 脑梗(非梗死) -> 脑梗死；防"脑梗死"变"脑梗死梗死"
    s = re.sub(r'脑梗(?!死)', '脑梗死', s)
    # 早搏 -> 期前收缩（防"期前收缩"变"期前期前收缩"）
    s = re.sub(r'(?<!期前)早搏', '期前收缩', s)
    s = s.lower().replace('（', '(').replace('）', ')')
    s = re.sub(r'[\s\u3000,，、。.．、；;:：\-_/\\\\"\'“”‘’（）()\[\]【】<>《》]+', '', s)
    out = []
    for i, ch in enumerate(s):
        nxt = s[i + 1] if i + 1 < len(s) else ''
        if ch in ROMAN:
            out.append(ROMAN[ch])
        elif ch in CN_NUM and nxt in '度级期型':
            out.append(CN_NUM[ch])
        else:
            out.append(ch)
    s = ''.join(out)
    for k in ('VIII', 'III', 'VII', 'II', 'IV', 'VI', 'V', 'I'):
        d = {'I': '1', 'II': '2', 'III': '3', 'IV': '4', 'V': '5',
             'VI': '6', 'VII': '7', 'VIII': '8'}[k]
        s = s.replace(k + '度', d + '度').replace(k + '级', d + '级')
    return s


def split_aliases(name):
    """从 ICD 名称中拆出 [别名]；返回 [主名, 别名...]"""
    main = re.sub(r'\[[^\]]*\]', '', name or '').strip()
    aliases = re.findall(r'\[([^\]]+)\]', name or '')
    parts = [main] + [a.strip() for a in aliases if a.strip()]
    return [p for p in parts if p]


def modifier_penalty(entry_name, query_norm):
    """条目名里出现、查询名没有的修饰词数量"""
    n = 0
    for m in MODIFIERS:
        if m in entry_name and m not in query_norm:
            n += 1
    return n

# ================================================================ 人工校正表

# 国临版2.0 中无对应/易误配的高频诊断名 -> (编码, 国临版标准名, 备注)
CORRECTIONS = {
    '肝功能异常': ('K72.905', '肝功能不全', '国临版无"肝功能异常"，取"肝功能不全"'),
    '糖耐量异常': ('E16.800x901', '糖耐量受损', '国临版无"糖耐量异常"，取"糖耐量受损"'),
    '急性肾损伤': ('N17.900', '急性肾衰竭', '国临版无"急性肾损伤"，取"急性肾衰竭"'),
    '急性肾损伤（1期）': ('N17.900', '急性肾衰竭', '同"急性肾损伤"'),
    '急性肾损伤（2期）': ('N17.900', '急性肾衰竭', '同"急性肾损伤"'),
    '卵圆孔未闭': ('Q21.101', '中央型房间隔缺损(卵圆孔型)', '国临版仅此编码'),
    '低高密度脂蛋白胆固醇血症': ('E78.600x006', '高密度脂蛋白缺乏', '低HDL≈高密度脂蛋白缺乏'),
    '双房增大': ('I51.700x003', '心房扩大', '双房=心房扩大'),
    '室间隔增厚': ('I51.706', '室间隔肥大', '国临版用"肥大"'),
    '室间隔基底段增厚': ('I51.706', '室间隔肥大', '国临版用"肥大"'),
    '频发室性早搏': ('I49.300x001', '频发性室性期前收缩', '早搏=期前收缩'),
    '血小板减少症': ('D69.400', '血小板减少，其他原发性的', '国临版无单纯"血小板减少症"'),
    '脂肪肝（非酒精性）': ('K76.000', '脂肪肝', '国临版无非酒精性脂肪肝独立码'),
    '非酒精性脂肪肝': ('K76.000', '脂肪肝', '国临版无非酒精性脂肪肝独立码'),
    '慢性心力衰竭急性加重': ('I50.908', '慢性心力衰竭', '国临版无"急性加重"'),
    '慢性心力衰竭急性失代偿': ('I50.908', '慢性心力衰竭', '国临版无"急性失代偿"'),
    '升主动脉增宽': ('I71.203', '升主动脉扩张', '增宽=扩张'),
    '主动脉增宽': ('I71.901', '主动脉扩张', '增宽=扩张'),
    '左房增大': ('I51.703', '左房扩大', '增大=扩大'),
    '左心房增大': ('I51.703', '左房扩大', '增大=扩大'),
    '右房增大': ('I51.704', '右房扩大', '增大=扩大'),
    '高血压病': ('I10.x09', '原发性高血压', '国临版"高血压病"仅分度编码，无分级取"原发性高血压"'),
    '高血压病1级（很高危）': ('I10.x00x024', '高血压病1级（极高危）', '国临版无"很高危"，取"极高危"'),
    '高血压病2级（很高危）': ('I10.x00x028', '高血压病2级（极高危）', '国临版无"很高危"，取"极高危"'),
    '高血压病3级（很高危）': ('I10.x00x032', '高血压病3级（极高危）', '国临版无"很高危"，取"极高危"'),
    '社区获得性肺炎': ('J15.902', '社区获得性肺炎，非重症', '无重症证据时按非重症'),
    '贫血（轻度）': ('D64.901', '轻度贫血', '程度词位置变体'),
    '贫血（中度）': ('D64.902', '中度贫血', '程度词位置变体'),
    '贫血（重度）': ('D64.903', '重度贫血', '程度词位置变体'),
    '糖代谢异常': ('E74.901', '糖代谢紊乱', '国临版用"糖代谢紊乱"'),
    '左肾囊肿': ('N28.101', '单纯性肾囊肿', ''),
    '右肾囊肿': ('N28.101', '单纯性肾囊肿', ''),
    '双肾囊肿': ('N28.101', '单纯性肾囊肿', ''),
    '肾囊肿': ('N28.101', '单纯性肾囊肿', ''),
    '胆囊胆固醇沉积': ('K82.400', '胆囊胆固醇沉着症', '国临版用"沉着症"'),
    '双肺感染': ('J98.414', '肺部感染', ''),
    '轻度肺气肿': ('J43.900', '肺气肿', '国临版无"轻度肺气肿"'),
    '脑梗死（陈旧性）': ('I69.300x002', '陈旧性脑梗死', ''),
    '急性肾损伤（3期）': ('N17.900', '急性肾衰竭', '同"急性肾损伤"'),
    '卵圆孔未闭（复杂型）': ('Q21.101', '中央型房间隔缺损(卵圆孔型)', '国临版仅此编码'),
    '高血糖状态': ('R73.900x001', '血糖升高', '无糖尿病语境的高血糖'),
    '左肾结石': ('N20.000', '肾结石', ''),
    '右肾结石': ('N20.000', '肾结石', ''),
    '肾结石': ('N20.000', '肾结石', ''),
    '肺大泡': ('J43.901', '肺大疱', '泡=疱'),
    '左心耳封堵术后': ('Z95.800x005', '心脏介入封堵术后状态', ''),
    '左心耳封堵术后状态': ('Z95.800x005', '心脏介入封堵术后状态', ''),
}

# ================================================================ 加载 ICD 库

def load_icd_lib(xlsx_path):
    lib_path = os.path.join(BASE, 'ICD库_国临版2.0.json')
    if os.path.exists(lib_path):
        with open(lib_path, encoding='utf-8') as f:
            return json.load(f)
    from openpyxl import load_workbook
    print('读取国临版2.0 xlsx: %s ...' % xlsx_path)
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        code, name = (row[0], row[1]) if len(row) >= 2 else (None, None)
        if not code or not name:
            continue
        code = str(code).strip()
        name = str(name).strip()
        parts = split_aliases(name)
        rows.append({'code': code, 'name': name,
                     'norm': norm(name), 'norms': [norm(p) for p in parts],
                     'main': parts[0]})
    wb.close()
    lib = {'schema': 'ICD-10 国临版2.0', 'count': len(rows), 'entries': rows,
           'exported_at': datetime.now().strftime('%Y-%m-%d %H:%M')}
    with open(lib_path, 'w', encoding='utf-8') as f:
        json.dump(lib, f, ensure_ascii=False, indent=1)
    print('[OK] ICD库_国临版2.0.json  %d 条' % len(rows))
    return lib


def build_index(lib):
    exact = defaultdict(list)
    by_first = defaultdict(list)
    for e in lib['entries']:
        for n in e['norms']:
            exact[n].append(e)
        if e['norm']:
            by_first[e['norm'][0]].append(e)
    return exact, by_first


def pick_exact(hits, query_norm):
    """精确命中优选：修饰词最少 -> 名称最短 -> 编码字典序"""
    hits.sort(key=lambda e: (modifier_penalty(e['name'], query_norm),
                             len(e['norm']), e['code']))
    return hits[0]


def pick_contains(entries, query_norm):
    """包含命中优选：长度差最小 -> 修饰词最少"""
    def score(e):
        en = e['norm']
        return (abs(len(en) - len(query_norm)),
                modifier_penalty(e['name'], query_norm),
                len(en))
    entries.sort(key=score)
    return entries[0]

# ================================================================ 匹配

def match_doc_name(dname, exact, by_first, entries):
    n = norm(dname)
    corr = CORRECTIONS.get(dname)
    if corr:
        return (corr[0], corr[1], 'correction')
    hits = exact.get(n, [])
    if hits:
        e = pick_exact(hits, n)
        return (e['code'], e['name'], 'exact')
    if len(n) >= 4:
        best = None
        for e in entries:
            en = e['norm']
            if not en or len(en) < 4:
                continue
            if len(en) >= len(n):
                if en.find(n) >= 0:
                    if best is None or (abs(len(en) - len(n)), modifier_penalty(e['name'], n)) < \
                                       (abs(len(best['norm']) - len(n)), modifier_penalty(best['name'], n)):
                        best = e
            elif len(n) >= len(en) + 2:
                if n.find(en) >= 0:
                    if best is None or (abs(len(en) - len(n)), modifier_penalty(e['name'], n)) < \
                                       (abs(len(best['norm']) - len(n)), modifier_penalty(best['name'], n)):
                        best = e
        if best:
            return (best['code'], best['name'], 'contains')
    bucket = by_first.get(n[0], [])
    if bucket:
        import difflib
        blist = [e['norm'] for e in bucket]
        m = difflib.get_close_matches(n, blist, n=1, cutoff=0.82)
        if m:
            e = bucket[blist.index(m[0])]
            return (e['code'], e['name'], 'fuzzy')
    return None


def match_ai_name(aname, exact, by_first, entries, doc_code_map):
    n = norm(aname)
    hits = exact.get(n, [])
    if hits:
        e = pick_exact(hits, n)
        return (e['code'], e['name'], 'icd_direct', 'high', '')
    code, icd_name = doc_code_map.get(aname, ('', ''))
    # 人工校正表（国临版无对应/易误配的高频名），优先于医生锚点与包含/模糊
    corr = CORRECTIONS.get(aname)
    if corr:
        return (corr[0], corr[1], 'correction', 'medium', corr[2])
    if code:
        return (code, icd_name, 'doctor_anchor', 'high', '')
    if len(n) >= 4:
        best = None
        for e in entries:
            en = e['norm']
            if not en or len(en) < 4:
                continue
            if len(en) >= len(n):
                if en.find(n) >= 0:
                    if best is None or abs(len(en) - len(n)) < abs(len(best['norm']) - len(n)):
                        best = e
            elif len(n) >= len(en) + 2:
                if n.find(en) >= 0:
                    if best is None or abs(len(en) - len(n)) < abs(len(best['norm']) - len(n)):
                        best = e
        if best:
            diff = abs(len(best['norm']) - len(n))
            conf = 'high' if diff <= 2 else 'low'
            return (best['code'], best['name'], 'contains', conf, '')
    bucket = by_first.get(n[0], [])
    if bucket:
        import difflib
        blist = [e['norm'] for e in bucket]
        m = difflib.get_close_matches(n, blist, n=1, cutoff=0.82)
        if m:
            e = bucket[blist.index(m[0])]
            return (e['code'], e['name'], 'fuzzy', 'low', '')
    return ('', '', 'unmapped', 'low', '')

# ================================================================ 主流程

def load_doctor_names():
    path = os.path.join(BASE, '医生标准名清单.json')
    if os.path.exists(path):
        with open(path, encoding='utf-8') as f:
            return [e['standard_name'] for e in json.load(f)['entries']]
    names, seen = [], set()
    with open(os.path.join(BASE, 'current_diagnoses.txt'), encoding='utf-8') as f:
        for line in f:
            parts = line.rstrip('\r\n').split('|#|', maxsplit=1)
            if len(parts) >= 2:
                t = parts[1].strip().replace('\\u000A', '\n')
                if t not in seen:
                    seen.add(t)
                    names.append(t)
    return names


def load_ai_names():
    """AI 诊断名 -> 频次（复用 parse_and_generate.py 的三层兼容解析，口径一致）"""
    import importlib.util
    spec = importlib.util.spec_from_file_location('pag', os.path.join(BASE, 'parse_and_generate.py'))
    pag = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(pag)
    ai = Counter()
    for rec in pag.load_records(os.path.join(BASE, 'diagnosis_analysis_records.txt')):
        blocks, _ = pag.parse_ai_diagnoses(rec['content'])
        for b in blocks:
            if b['name']:
                ai[b['name'].strip()] += 1
    return ai


def main():
    lib_path = os.path.join(BASE, 'ICD库_国临版2.0.json')
    if len(sys.argv) >= 2:
        lib = load_icd_lib(sys.argv[1])
    elif os.path.exists(lib_path):
        with open(lib_path, encoding='utf-8') as f:
            lib = json.load(f)
        print('复用现有 ICD 库: %s（%d 条）' % (lib_path, lib['count']))
    else:
        print('[ERR] 首次运行需提供国临版2.0 xlsx 路径: python build_icd_dict.py <xlsx>')
        sys.exit(1)
    entries = lib['entries']
    exact, by_first = build_index(lib)
    print('ICD 库 %d 条，精确索引 %d 键' % (lib['count'], len(exact)))

    doc_names = load_doctor_names()
    print('医生标准名 %d 个' % len(doc_names))
    doc_code_map, doc_unmapped = {}, []
    for dn in doc_names:
        hit = match_doc_name(dn, exact, by_first, entries)
        if hit:
            doc_code_map[dn] = (hit[0], hit[1])
        else:
            doc_unmapped.append(dn)
    print('医生名映射成功 %d / %d' % (len(doc_code_map), len(doc_names)))

    ai_names = load_ai_names()
    print('AI 诊断名 %d 个' % len(ai_names))
    out_entries, unmapped = [], []
    by_src = Counter()
    by_conf = Counter()
    for aname, freq in ai_names.most_common():
        code, icd_name, src, conf, note = match_ai_name(aname, exact, by_first, entries, doc_code_map)
        by_src[src] += 1
        by_conf[conf] += 1
        if src == 'unmapped':
            unmapped.append({'ai_name': aname, 'freq': freq})
        out_entries.append({'ai_name': aname, 'freq': freq, 'icd10_code': code,
                            'icd_name': icd_name, 'match_source': src,
                            'confidence': conf, 'note': note})

    mapped_items = sum(e['freq'] for e in out_entries if e['icd10_code'])
    ai_items = sum(ai_names.values())
    out = {
        'schema_version': '3.0',
        'purpose': 'AI诊断名 → ICD-10国临版2.0 标准名/编码（Prompt模板优化 + 合规判定）',
        'built_at': datetime.now().strftime('%Y-%m-%d %H:%M'),
        'icd_lib': 'ICD库_国临版2.0.json',
        'stats': {'ai_names': len(out_entries), 'ai_items_total': ai_items,
                  'mapped': sum(by_src[s] for s in ('icd_direct', 'doctor_anchor', 'contains', 'fuzzy', 'correction')),
                  'unmapped': by_src['unmapped'],
                  'by_source': dict(by_src), 'by_confidence': dict(by_conf),
                  'mapped_items': mapped_items,
                  'mapped_item_ratio': round(100.0 * mapped_items / ai_items, 1)},
        'entries': out_entries,
    }
    with open(os.path.join(BASE, '诊断编码词典.json'), 'w', encoding='utf-8') as f:
        json.dump(out, f, ensure_ascii=False, indent=1)
    print('[OK] 诊断编码词典.json  mapped=%d unmapped=%d 覆盖AI条数 %.1f%%'
          % (out['stats']['mapped'], out['stats']['unmapped'], out['stats']['mapped_item_ratio']))

    gen_unmapped_xlsx(unmapped, doc_unmapped)
    gen_report(out, doc_unmapped, unmapped)


def gen_unmapped_xlsx(unmapped, doc_unmapped):
    from openpyxl import Workbook
    from openpyxl.styles import Font
    wb = Workbook()
    ws = wb.active
    ws.title = '未映射清单'
    ws.append(['AI诊断名', '频次', 'ICD-10编码(人工补)', '国临版标准名(人工补)', '备注'])
    for c in ws[1]:
        c.font = Font(bold=True)
    for e in sorted(unmapped, key=lambda x: -x['freq']):
        ws.append([e['ai_name'], e['freq'], '', '', ''])
    for i, w in enumerate([44, 8, 18, 44, 24], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws2 = wb.create_sheet('医生名未映射')
    ws2.append(['医生标准名', 'ICD-10编码(人工补)'])
    for c in ws2[1]:
        c.font = Font(bold=True)
    for dn in doc_unmapped:
        ws2.append([dn, ''])
    for i, w in enumerate([50, 18], 1):
        ws2.column_dimensions[chr(64 + i)].width = w
    path = os.path.join(BASE, '诊断编码词典_未映射清单.xlsx')
    wb.save(path)
    print('[OK] %s  (未映射 %d 条, 医生未映射 %d 条)' % (path, len(unmapped), len(doc_unmapped)))


def gen_report(out, doc_unmapped, unmapped):
    L = []
    s = out['stats']
    L.append('# 诊断编码词典构建报告（ICD-10 国临版2.0 全量筛查）\n')
    L.append('- 生成时间：%s' % out['built_at'])
    L.append('- ICD 库：%s' % out['icd_lib'])
    L.append('- AI 诊断名 %d 个 / 覆盖 AI 条数 %d\n' % (s['ai_names'], s['ai_items_total']))
    L.append('## 一、映射统计\n')
    L.append('| 指标 | 数量 |')
    L.append('|------|------|')
    L.append('| 已映射 | %d（%.1f%%） |' % (s['mapped'], 100.0 * s['mapped'] / s['ai_names']))
    L.append('| 未映射（需人工补码） | %d |' % s['unmapped'])
    L.append('| 覆盖 AI 条数占比 | %.1f%% |' % s['mapped_item_ratio'])
    L.append('')
    L.append('| 匹配来源 | 条数 |')
    L.append('|----------|------|')
    for k, v in sorted(s['by_source'].items(), key=lambda kv: -kv[1]):
        L.append('| %s | %d |' % (k, v))
    L.append('')
    L.append('| 置信度 | 条数 |')
    L.append('|--------|------|')
    for k, v in sorted(s['by_confidence'].items(), key=lambda kv: -kv[1]):
        L.append('| %s | %d |' % (k, v))
    L.append('')
    L.append('## 二、未映射 AI 诊断名 Top60（AI 命名偏离 ICD 编码体系的核心集合）\n')
    L.append('> 含大量检验结果型/影像描述型表述（如"糖化血红蛋白升高""左房增大"），'
             '本就不应作为诊断名输出——模板优化方向：禁止把检查指标/影像发现直接当诊断名。\n')
    L.append('| AI诊断名 | 频次 |')
    L.append('|----------|------|')
    for e in sorted(unmapped, key=lambda x: -x['freq'])[:60]:
        L.append('| %s | %d |' % (e['ai_name'], e['freq']))
    L.append('')
    L.append('## 三、医生标准名未映射（%d 个）\n' % len(doc_unmapped))
    for dn in sorted(doc_unmapped):
        L.append('- %s' % dn)
    L.append('')

    inj = [e for e in out['entries'] if e['icd10_code']]
    inj.sort(key=lambda e: (-e['freq'], e['confidence'] != 'high'))
    top = inj[:150]
    cov = sum(e['freq'] for e in inj) / s['ai_items_total']
    top_cov = sum(e['freq'] for e in top) / s['ai_items_total']
    L.append('## 四、Prompt 注入清单（高频 AI 诊断名 → 国临版标准名+编码）\n')
    L.append('> 全量已映射 %d 个名，覆盖 AI 条数 %.1f%%；Top%d 名覆盖 %.1f%%。'
             % (len(inj), 100.0 * cov, len(top), 100.0 * top_cov))
    L.append('> 注入方式：作为诊断分析模板"命名规范"段附录——要求 AI 对以下诊断优先采用标准表述；'
             '未列出的诊断名要求遵循 ICD-10 国临版术语，禁止用检验/影像描述替代诊断名。\n')
    L.append('| AI 常用名 | 频次 | 标准名（国临版2.0） | ICD编码 | 置信 |')
    L.append('|-----------|------|----------------------|---------|------|')
    for e in top:
        L.append('| %s | %d | %s | %s | %s |'
                 % (e['ai_name'], e['freq'], e['icd_name'], e['icd10_code'], e['confidence']))
    L.append('')
    path = os.path.join(BASE, '诊断编码词典_统计报告.md')
    with open(path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(L))
    print('[OK] %s' % path)


from collections import defaultdict

if __name__ == '__main__':
    main()
