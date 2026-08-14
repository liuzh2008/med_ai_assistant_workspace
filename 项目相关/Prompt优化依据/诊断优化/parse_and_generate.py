# -*- coding: utf-8 -*-
"""
诊断分析对比文档生成器

输入：extract_02_results.sql / extract_03_diagnoses.sql 的导出文件
      - diagnosis_analysis_records.txt   (AI诊断 + 依据原文，每行一条记录)
      - current_diagnoses.txt            (目前诊断，每行一条)
输出：
      - 诊断分析对比报告_YYYYMMDD.md      (Markdown 阅读版)
      - 诊断分析对比明细_YYYYMMDD.xlsx    (Excel 明细 + 聚合统计)

判定逻辑（沿用《AI诊断数据采集方案》Levenshtein 阈值）：
      相似度 = 1.0        -> 直接采纳
      0.45 <= 相似度 < 1.0 -> 被修改
      相似度 < 0.45       -> 未采纳（AI 侧）/ 医生新增（医生侧）

用法：python parse_and_generate.py
"""
import os
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime

SEP = '|#|'
SEP_ESCAPED = '\u00a6#\u00a6'  # ¦#¦

SIM_MODIFIED = 0.45  # 0.45~1.0 视为"被修改"（与6月方案一致）


# ---------------------------------------------------------------- 解析工具

def unescape(s):
    return (s.replace('\\u000A', '\n')
             .replace('\\u0009', '\t')
             .replace(SEP_ESCAPED, SEP))


def strip_thinking(content):
    return re.sub(r'<thinking>[\s\S]*?</thinking>', '', content or '', flags=re.I)


def extract_field(field, text):
    m = re.search(r'(?:#{3,4}\s*)?%s[:：]\s*([\s\S]*?)(?=#{3,4}|$)' % re.escape(field),
                  text, re.I)
    return m.group(1).strip() if m else ''


def parse_block(bc):
    b = {'index': '', 'name': '', 'category': '', 'basis': '',
         'differential': '', 'supplement': ''}
    m = re.search(r'(?:#{3,4}\s*)?诊断编号[:：]\s*(\d+)', bc, re.I)
    if m:
        b['index'] = m.group(1)
    b['name'] = extract_field('诊断名称', bc)
    b['category'] = extract_field('诊断类别', bc)
    b['basis'] = extract_field('诊断依据', bc)
    b['differential'] = extract_field('鉴别诊断', bc)
    b['supplement'] = extract_field('补充说明', bc)
    if not b['name']:
        m = re.search(r'(?:#{3,4}\s*)?诊断[:：]\s*(.+?)(?:\n|$)', bc, re.I)
        if m:
            b['name'] = m.group(1).strip()
    return b


def extract_diagnosis_blocks(content):
    """标准格式：### 诊断列表 区块内的诊断块（与前端 extractDiagnosisBlocks 一致）"""
    content = strip_thinking(content)
    m = re.search(r'###\s*诊断列表\s*\n([\s\S]*?)(?=\n###\s[^#]|$)', content)
    if not m:
        return []
    lst = m.group(1)
    starts = []
    for mm in re.finditer(r'(?:#{3,4}\s*诊断编号[:：])|(?:\n(?=#{3,4}\s*诊断名称[:：]))', lst):
        starts.append(mm.start())
    if not starts:
        for mm in re.finditer(r'#{3,4}\s*诊断名称[:：]', lst):
            starts.append(mm.start())
    blocks = []
    for i, st in enumerate(starts):
        en = starts[i + 1] if i + 1 < len(starts) else len(lst)
        bc = lst[st:en].strip()
        if not bc:
            continue
        b = parse_block(bc)
        if b and b['name']:
            blocks.append(b)
    return blocks


def extract_diagnosis_names(content):
    """格式B：诊断名称: xxx 标记（无完整块结构）"""
    content = strip_thinking(content)
    names = []
    for m in re.finditer(r'(?:#{2,4}\s*)?诊断名称[:：]\s*(.+?)(?:\n|$)', content):
        names.append(m.group(1).strip())
    if not names:
        for m in re.finditer(r'诊断[:：]\s*(.+?)(?:\n|$)', content):
            names.append(m.group(1).strip())
    return list(dict.fromkeys([n for n in names if n]))


def extract_legacy_names(content):
    """格式C（旧版模板）：1. **诊断名**: ... 编号加粗列表"""
    content = strip_thinking(content)
    names = []
    for m in re.finditer(r'\d+[\.、]\s*\*\*([^*\n]{2,60})\*\*[:：]', content):
        names.append(m.group(1).strip())
    return list(dict.fromkeys([n for n in names if n]))


def parse_ai_diagnoses(content):
    """三层兼容解析：标准块 -> 名称标记 -> 旧版编号加粗，返回诊断块列表"""
    blocks = extract_diagnosis_blocks(content)
    if blocks:
        return blocks, '标准格式'
    names = extract_diagnosis_names(content)
    if names:
        return [{'index': '', 'name': n, 'category': '', 'basis': '',
                 'differential': '', 'supplement': ''} for n in names], '名称标记'
    names = extract_legacy_names(content)
    if names:
        return [{'index': '', 'name': n, 'category': '', 'basis': '',
                 'differential': '', 'supplement': ''} for n in names], '旧版格式'
    return [], '解析失败'


def normalize(s):
    return re.sub(r'[\s\u3000,，、。.、；;：:（）()\-_/]', '', s or '').lower()


def levenshtein_sim(a, b):
    sa, sb = normalize(a), normalize(b)
    if not sa and not sb:
        return 1.0
    if not sa or not sb:
        return 0.0
    m, n = len(sa), len(sb)
    prev = list(range(n + 1))
    for i in range(1, m + 1):
        cur = [i] + [0] * n
        for j in range(1, n + 1):
            cur[j] = min(prev[j] + 1, cur[j - 1] + 1,
                         prev[j - 1] + (0 if sa[i - 1] == sb[j - 1] else 1))
        prev = cur
    return 1 - prev[n] / max(m, n)


# ---------------------------------------------------------------- 数据加载

def load_records(path):
    recs = []
    with open(path, encoding='utf-8') as f:
        for line in f:
            line = line.rstrip('\r\n')
            if not line:
                continue
            parts = line.split(SEP, maxsplit=5)
            if len(parts) < 6:
                print('[WARN] 记录字段不足，跳过: %s' % line[:80])
                continue
            recs.append({'result_id': parts[0].strip(), 'template': parts[1].strip(),
                         'patient_hash': parts[2].strip(), 'exec_time': parts[3].strip(),
                         'status': parts[4].strip(), 'content': unescape(parts[5])})
    return recs


def load_diagnoses(path):
    by_patient = defaultdict(list)
    with open(path, encoding='utf-8') as f:
        for line in f:
            line = line.rstrip('\r\n')
            if not line:
                continue
            parts = line.split(SEP, maxsplit=5)
            if len(parts) < 6:
                continue
            by_patient[parts[0].strip()].append({
                'text': unescape(parts[1].strip()),
                'is_primary': parts[2].strip() == '1'})
    return by_patient


# ---------------------------------------------------------------- 配对判定

def classify(records, by_patient):
    """对每条诊断分析记录产出对比行"""
    rows = []      # AI 侧 + 医生新增侧明细
    for rec in records:
        blocks, fmt = parse_ai_diagnoses(rec['content'])
        doc_diags = by_patient.get(rec['patient_hash'], [])
        base = {'patient': rec['patient_hash'], 'time': rec['exec_time'],
                'template': rec['template'], 'format': fmt}

        if not blocks:
            rows.append(dict(base, kind='解析失败', ai_name='', ai_category='',
                             verdict='解析失败', doc_name='', sim=None,
                             basis='', differential='', supplement='',
                             doc_diags='、'.join(d['text'] for d in doc_diags)))
            continue

        # AI 侧：每个 AI 诊断 vs 目前诊断
        for b in blocks:
            best_sim, best_name = 0.0, ''
            for d in doc_diags:
                s = levenshtein_sim(b['name'], d['text'])
                if s > best_sim:
                    best_sim, best_name = s, d['text']
            if not doc_diags:
                verdict, best_name, best_sim = '医生无诊断', '', None
            elif best_sim >= 0.999:
                verdict = '直接采纳'
            elif best_sim >= SIM_MODIFIED:
                verdict = '被修改'
            else:
                verdict = '未采纳'
            rows.append(dict(base, kind='AI', ai_name=b['name'],
                             ai_category=b['category'], verdict=verdict,
                             doc_name=best_name,
                             sim=round(best_sim, 3) if best_sim else None,
                             basis=b['basis'], differential=b['differential'],
                             supplement=b['supplement'], doc_diags=''))

        # 医生侧：医生诊断里 AI 完全没覆盖到的 -> 医生新增
        ai_names = [b['name'] for b in blocks]
        for d in doc_diags:
            best = max((levenshtein_sim(n, d['text']) for n in ai_names), default=0.0)
            if best < SIM_MODIFIED:
                rows.append(dict(base, kind='医生', ai_name='', ai_category='',
                                 verdict='医生新增', doc_name=d['text'],
                                 sim=round(best, 3) if ai_names else None,
                                 basis='', differential='', supplement='',
                                 doc_diags=''))
    return rows


# ---------------------------------------------------------------- 报告生成

def brief(text, limit=90):
    t = re.sub(r'\s+', ' ', text or '').strip()
    return t if len(t) <= limit else t[:limit] + '…'


def gen_markdown(records, rows, out_path):
    n_rec = len(records)
    n_patients = len(set(r['patient_hash'] for r in records))
    ai_rows = [r for r in rows if r['kind'] == 'AI']
    doc_rows = [r for r in rows if r['kind'] == '医生']
    failed = [r for r in rows if r['kind'] == '解析失败']
    vc = Counter(r['verdict'] for r in ai_rows)
    drc = Counter(r['verdict'] for r in doc_rows)

    L = []
    L.append('# 诊断分析 AI vs 目前诊断 对比报告\n')
    L.append('- 生成时间：%s' % datetime.now().strftime('%Y-%m-%d %H:%M'))
    L.append('- 数据范围：近 90 天（提取脚本口径，SYSDATE-90）\n')
    L.append('## 一、概览\n')
    L.append('| 指标 | 数量 |')
    L.append('|------|------|')
    L.append('| 诊断分析记录数 | %d |' % n_rec)
    L.append('| 涉及患者数 | %d |' % n_patients)
    L.append('| AI 诊断条数 | %d |' % len(ai_rows))
    L.append('| 解析失败记录数（AI 输出非标准格式） | %d |' % len(failed))
    L.append('| 目前诊断条数（医生侧） | %d |' % len(doc_rows))
    L.append('')
    L.append('| AI 诊断判定 | 条数 | 占比 |')
    L.append('|------------|------|------|')
    total = len(ai_rows) or 1
    for v in ('直接采纳', '被修改', '未采纳', '医生无诊断'):
        c = vc.get(v, 0)
        L.append('| %s | %d | %.1f%% |' % (v, c, c * 100.0 / total))
    L.append('')
    L.append('医生侧新增诊断 %d 条次（AI 完全未提及，为模板遗漏线索；同一患者多次分析会累计）。\n' % drc.get('医生新增', 0))

    # 二、模板缺陷线索聚合
    L.append('## 二、模板缺陷线索（按 AI 诊断名聚合）\n')
    agg = defaultdict(lambda: {'total': 0, 'adopted': 0, 'modified': 0,
                               'rejected': 0, 'doc_names': Counter()})
    for r in ai_rows:
        a = agg[r['ai_name']]
        a['total'] += 1
        if r['verdict'] == '直接采纳':
            a['adopted'] += 1
        elif r['verdict'] == '被修改':
            a['modified'] += 1
        elif r['verdict'] == '未采纳':
            a['rejected'] += 1
        if r['doc_name']:
            a['doc_names'][r['doc_name']] += 1
    ordered = sorted(agg.items(),
                     key=lambda kv: (-(kv[1]['rejected'] + kv[1]['modified']), -kv[1]['total']))
    L.append('| AI诊断名 | 出现 | 采纳 | 修改 | 未采纳 | 修改+未采纳率 | 常见医生版本 |')
    L.append('|----------|------|------|------|--------|---------------|--------------|')
    for name, a in ordered[:40]:
        bad = a['modified'] + a['rejected']
        rate = bad * 100.0 / a['total'] if a['total'] else 0
        doc_ver = '；'.join('%s×%d' % kv for kv in a['doc_names'].most_common(2))
        L.append('| %s | %d | %d | %d | %d | %.0f%% | %s |'
                 % (name, a['total'], a['adopted'], a['modified'], a['rejected'], rate, doc_ver))
    L.append('')

    # 三、被修改明细（价值最高）
    mod = [r for r in ai_rows if r['verdict'] == '被修改']
    L.append('## 三、被修改明细（AI 方向对但命名不精确，最能定位模板缺陷）\n')
    if mod:
        mod.sort(key=lambda r: (normalize(r['ai_name']), r['sim'] or 0))
        cur_name = None
        for r in mod:
            if normalize(r['ai_name']) != cur_name:
                cur_name = normalize(r['ai_name'])
                L.append('### %s → 医生版本\n' % r['ai_name'])
            doc_show = r['doc_name'] or '（无匹配）'
            L.append('- 患者`%s`（%s）：医生最终 **%s**，相似度 %.2f'
                     % (r['patient'], r['time'], doc_show, r['sim'] or 0))
            if r['basis']:
                L.append('  - AI依据：%s' % brief(r['basis']))
        L.append('')
    else:
        L.append('（无）\n')

    # 四、未采纳明细
    rej = [r for r in ai_rows if r['verdict'] == '未采纳']
    L.append('## 四、未采纳明细（AI 多写/无关诊断，反映模板冗余）\n')
    if rej:
        rej.sort(key=lambda r: (normalize(r['ai_name']), r['sim'] or 0))
        cur_name = None
        for r in rej:
            if normalize(r['ai_name']) != cur_name:
                cur_name = normalize(r['ai_name'])
                L.append('### %s\n' % r['ai_name'])
            doc_show = r['doc_name'] or '（无匹配）'
            L.append('- 患者`%s`（%s）：最接近的医生诊断 **%s**，相似度 %.2f'
                     % (r['patient'], r['time'], doc_show, r['sim'] or 0))
            if r['basis']:
                L.append('  - AI依据：%s' % brief(r['basis']))
        L.append('')
    else:
        L.append('（无）\n')

    # 五、医生新增（AI 遗漏）
    L.append('## 五、医生新增诊断（AI 完全未提及，模板遗漏线索）\n')
    if doc_rows:
        c = Counter(r['doc_name'] for r in doc_rows)
        L.append('| 医生诊断名 | 次数 |')
        L.append('|------------|------|')
        for name, cnt in c.most_common(30):
            L.append('| %s | %d |' % (name, cnt))
        L.append('')
    else:
        L.append('（无）\n')

    # 六、直接采纳清单（简表）
    ad = [r for r in ai_rows if r['verdict'] == '直接采纳']
    L.append('## 六、直接采纳清单（简表）\n')
    if ad:
        L.append('| 患者 | 时间 | AI诊断（采纳） |')
        L.append('|------|------|----------------|')
        for r in ad:
            L.append('| `%s` | %s | %s |' % (r['patient'], r['time'], r['ai_name']))
        L.append('')
    else:
        L.append('（无）\n')

    # 七、解析失败记录
    if failed:
        L.append('## 七、解析失败记录（AI 输出非标准格式，建议人工查看）\n')
        for r in failed:
            L.append('- 患者`%s`（%s）：目前诊断：%s'
                     % (r['patient'], r['time'], brief(r['doc_diags'], 200)))
        L.append('')

    with open(out_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(L))
    print('[OK] Markdown 报告: %s' % out_path)


def gen_excel(records, rows, out_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = '对比明细'
    headers = ['患者哈希', '执行时间', '模板', 'AI诊断名', 'AI类别', '判定',
               '医生诊断名(最匹配)', '相似度', 'AI诊断依据', '鉴别诊断', '补充说明']
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in rows:
        ws.append([r['patient'], r['time'], r['template'], r['ai_name'],
                   r['ai_category'], r['verdict'], r['doc_name'],
                   r['sim'] if r['sim'] is not None else '',
                   r['basis'], r['differential'], r['supplement']])
    widths = [18, 20, 10, 30, 10, 10, 36, 8, 60, 40, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    # 聚合统计 sheet
    ws2 = wb.create_sheet('AI诊断聚合')
    agg = defaultdict(lambda: {'total': 0, 'adopted': 0, 'modified': 0,
                               'rejected': 0, 'doc_names': Counter()})
    for r in rows:
        if r['kind'] != 'AI':
            continue
        a = agg[r['ai_name']]
        a['total'] += 1
        if r['verdict'] == '直接采纳':
            a['adopted'] += 1
        elif r['verdict'] == '被修改':
            a['modified'] += 1
        elif r['verdict'] == '未采纳':
            a['rejected'] += 1
        if r['doc_name']:
            a['doc_names'][r['doc_name']] += 1
    ws2.append(['AI诊断名', '出现次数', '直接采纳', '被修改', '未采纳',
                '修改+未采纳率', '常见医生版本'])
    for c in ws2[1]:
        c.font = Font(bold=True)
    for name, a in sorted(agg.items(), key=lambda kv: -kv[1]['total']):
        bad = a['modified'] + a['rejected']
        rate = bad * 100.0 / a['total'] if a['total'] else 0
        ws2.append([name, a['total'], a['adopted'], a['modified'], a['rejected'],
                    '%.0f%%' % rate,
                    '；'.join('%s×%d' % kv for kv in a['doc_names'].most_common(3))])
    for i, w in enumerate([30, 8, 8, 8, 8, 12, 60], 1):
        ws2.column_dimensions[chr(64 + i)].width = w

    wb.save(out_path)
    print('[OK] Excel 明细: %s' % out_path)


def main():
    base = os.path.dirname(os.path.abspath(__file__))
    rec_path = os.path.join(base, 'diagnosis_analysis_records.txt')
    diag_path = os.path.join(base, 'current_diagnoses.txt')
    if not os.path.exists(rec_path):
        print('[ERR] 缺少输入文件: %s' % rec_path)
        sys.exit(1)
    if not os.path.exists(diag_path):
        print('[ERR] 缺少输入文件: %s' % diag_path)
        sys.exit(1)

    print('读取诊断分析记录...')
    records = load_records(rec_path)
    print('读取目前诊断...')
    by_patient = load_diagnoses(diag_path)
    print('记录 %d 条 / 有诊断的患者 %d 位' % (len(records), len(by_patient)))

    rows = classify(records, by_patient)
    stamp = datetime.now().strftime('%Y%m%d')
    gen_markdown(records, rows, os.path.join(base, '诊断分析对比报告_%s.md' % stamp))
    gen_excel(records, rows, os.path.join(base, '诊断分析对比明细_%s.xlsx' % stamp))
    print('完成。')


if __name__ == '__main__':
    main()
