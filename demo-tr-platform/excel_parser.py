#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
教研室Excel数据解析脚本

功能：读取xlsx文件，处理合并单元格，输出JSON结构化数据到stdout。
用法：python excel_parser.py --file <xlsx_path> [--sheet <sheet_name>]
输出：JSON数组 [{字段名: 值, ...}, ...]
"""

import argparse
import json
import sys
from pathlib import Path

# 强制 stdout/stderr 输出 UTF-8 字节，避免 Windows 编码问题（BOM/UTF-16 等）
def _write_stdout(data):
    """将 dict/list 以 UTF-8 JSON 写入 stdout，追加换行。"""
    json_bytes = json.dumps(data, ensure_ascii=False, indent=2).encode('utf-8')
    sys.stdout.buffer.write(json_bytes + b'\n')
    sys.stdout.buffer.flush()

def _write_stderr(data):
    """将 dict 以 UTF-8 JSON 写入 stderr。"""
    json_bytes = json.dumps(data, ensure_ascii=False).encode('utf-8')
    sys.stderr.buffer.write(json_bytes + b'\n')
    sys.stderr.buffer.flush()

try:
    import openpyxl
except ImportError:
    _write_stderr({"error": "缺少 openpyxl 依赖，请执行: pip install openpyxl"})
    sys.exit(1)


# 列名中文 → 英文字段名映射（教改论文Sheet）
COLUMN_MAPPING = {
    "序号": "seq",
    "题目": "title",
    "项目": "title",
    "第一作者": "first_author",
    "教研室": "department",
    "通讯作者": "corresponding_author",
    "发表期刊": "journal",
    "期刊级别": "journal_level",
    "期刊级别(核心/科技核心/一般)": "journal_level",
    "时间": "pub_year",
    "备注": "remark",
}


def build_merged_cell_map(ws):
    """
    构建合并单元格映射表：key=(row, col) → value=左上角单元格的值。
    遍历所有合并单元格范围，将区域内每个坐标映射到左上角值。
    这样读取时遇到合并单元格就能直接获取正确的值，无需修改只读的MergedCell。
    """
    merged_map = {}
    for merged_range in ws.merged_cells.ranges:
        min_row = merged_range.min_row
        max_row = merged_range.max_row
        min_col = merged_range.min_col
        max_col = merged_range.max_col

        # 获取左上角单元格的值
        top_left_cell = ws.cell(row=min_row, column=min_col)
        top_left_value = top_left_cell.value

        # 将区域内所有坐标映射到左上角值
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                if row != min_row or col != min_col:
                    merged_map[(row, col)] = top_left_value
    return merged_map


def get_cell_value(ws, row, col, merged_map):
    """
    安全获取单元格值：优先检查合并单元格映射表，再读取实际单元格。
    """
    key = (row, col)
    if key in merged_map:
        return merged_map[key]
    return ws.cell(row=row, column=col).value


def normalize_header(header_text):
    """清理表头文本：去除换行、多余空格"""
    if header_text is None:
        return ""
    text = str(header_text).strip()
    # 替换换行为空格
    text = text.replace('\n', ' ').replace('\r', ' ')
    # 合并多余空格
    text = ' '.join(text.split())
    return text


def map_column_name(chinese_name):
    """将中文列名映射为英文字段名"""
    name = normalize_header(chinese_name)
    # 精确匹配
    if name in COLUMN_MAPPING:
        return COLUMN_MAPPING[name]
    # 模糊匹配：检查是否包含已知列名
    for key, value in COLUMN_MAPPING.items():
        if key in name or name in key:
            return value
    # 无法映射，返回清理后的中文名作为兜底
    return name.replace('(', '_').replace(')', '_').replace('/', '_').strip('_')


def is_data_row(row_values):
    """判断是否为有效数据行（至少有一个非空值，且不是纯表头/汇总行）"""
    non_empty = [v for v in row_values if v is not None and str(v).strip() != '']
    if len(non_empty) == 0:
        return False
    # 跳过可能只含序号或全是空的行
    combined = ' '.join(str(v) for v in non_empty)
    if len(non_empty) <= 1 and combined.strip().isdigit():
        return False
    return True


def find_header_row(ws, merged_map):
    """
    智能检测表头行：从第1行开始扫描，找到包含最多非空列的行作为表头。
    跳过纯标题行（合并单元格占整行且内容为标题文本）。
    """
    best_row = 1
    best_count = 0
    for row_idx in range(1, min(ws.max_row + 1, 5)):
        non_empty = 0
        for col_idx in range(1, min(ws.max_column + 1, 30)):
            val = get_cell_value(ws, row_idx, col_idx, merged_map)
            if val is not None and str(val).strip() != '':
                non_empty += 1
        if non_empty > best_count:
            best_count = non_empty
            best_row = row_idx
    return best_row


def find_effective_columns(ws, header_row, merged_map):
    """
    确定有效列数：从表头行向右扫描，找到最后一个有内容的列。
    同时检查数据行确保不是表头后的空列。
    """
    max_col = min(ws.max_column, 30)  # 最多检查30列
    effective_col = 0
    for col in range(1, max_col + 1):
        header_val = get_cell_value(ws, header_row, col, merged_map)
        if header_val is not None and str(header_val).strip() != '':
            effective_col = col
    return effective_col


def parse_excel(file_path, sheet_name=None, header_row=None):
    """
    解析Excel文件，返回结构化数据列表。

    Args:
        file_path: Excel文件路径
        sheet_name: 目标Sheet名称，为None时取第一个Sheet
        header_row: 表头行号（1-based），为None时自动检测

    Returns:
        list[dict]: 结构化数据列表
    """
    file_path = Path(file_path)
    if not file_path.exists():
        raise FileNotFoundError(f"文件不存在: {file_path}")

    wb = openpyxl.load_workbook(str(file_path), data_only=True)

    # 选择Sheet
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            # 尝试模糊匹配
            matched = [s for s in wb.sheetnames if sheet_name in s]
            if matched:
                sheet_name = matched[0]
            else:
                available = ', '.join(wb.sheetnames)
                raise ValueError(f"找不到Sheet '{sheet_name}'，可用的Sheet: {available}")
        ws = wb[sheet_name]
    else:
        ws = wb.active

    # 构建合并单元格映射表
    merged_map = build_merged_cell_map(ws)

    # 确定表头行
    if header_row is None:
        header_row = find_header_row(ws, merged_map)

    # 确定有效列数
    actual_max_col = find_effective_columns(ws, header_row, merged_map)

    # 读取表头
    headers = []
    for col in range(1, actual_max_col + 1):
        cell_value = get_cell_value(ws, header_row, col, merged_map)
        headers.append(normalize_header(cell_value))

    # 映射列名
    field_names = [map_column_name(h) for h in headers]

    # 读取数据行（从header_row+1开始）
    results = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        row_values = []
        for col_idx in range(1, actual_max_col + 1):
            value = get_cell_value(ws, row_idx, col_idx, merged_map)
            if value is not None:
                value = str(value).strip()
            row_values.append(value)

        if not is_data_row(row_values):
            continue

        # 构建字典
        row_dict = {}
        for i, field_name in enumerate(field_names):
            if i < len(row_values):
                val = row_values[i]
                if val == '' or val is None:
                    val = None
                row_dict[field_name] = val

        # 跳过完全空的行（所有字段都是None）
        non_none = sum(1 for v in row_dict.values() if v is not None)
        if non_none <= 1:
            continue

        # 类型转换：pub_year尝试转为整数
        if 'pub_year' in row_dict and row_dict['pub_year'] is not None:
            try:
                row_dict['pub_year'] = int(float(row_dict['pub_year']))
            except (ValueError, TypeError):
                pass

        results.append(row_dict)

    wb.close()
    return results


def main():
    parser = argparse.ArgumentParser(description='教研室Excel数据解析工具')
    parser.add_argument('--file', required=True, help='Excel文件路径 (.xlsx)')
    parser.add_argument('--sheet', default=None, help='目标Sheet名称（默认第一个Sheet）')
    parser.add_argument('--header-row', type=int, default=None, help='表头行号（1-based，默认自动检测）')
    parser.add_argument('--list-sheets', action='store_true', help='列出所有Sheet名称')

    args = parser.parse_args()

    try:
        if args.list_sheets:
            wb = openpyxl.load_workbook(args.file, data_only=True)
            _write_stdout({"sheets": wb.sheetnames})
            wb.close()
            return

        data = parse_excel(args.file, args.sheet, args.header_row)
        _write_stdout(data)

    except FileNotFoundError as e:
        _write_stderr({"error": str(e)})
        sys.exit(1)
    except ValueError as e:
        _write_stderr({"error": str(e)})
        sys.exit(1)
    except Exception as e:
        _write_stderr({"error": f"解析失败: {str(e)}"})
        sys.exit(1)


if __name__ == '__main__':
    main()
